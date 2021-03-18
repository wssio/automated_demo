# -* coding:utf-8 -*-
"""
time   :  2021/3/07 17:53
author :  zhangzhilong
effect :  读取Excel文件并转换为列表
"""

# from openpyxl import workbook
# workbook()新建表格
# wb = workbook()
from openpyxl import load_workbook


class excel:
    @staticmethod
    def read_excel(path):
        # 定义空列表
        data = []
        # load_workbook(文件名)打开已有表格，[工作簿名]获取工作簿
        excel_data = load_workbook(path[0])[path[1]]
        # range(x,y)从x到y-1的序列; .max_row获取工作簿的行数
        for i in range(1, int(str(excel_data.max_row)) + 1):
            # 排除表头
            if i != 1:
                # 向列表中添加一个空值
                data.append([])
                # .max_column获取工作簿的列数
                for j in range(1, int(str(excel_data.max_column)) + 1):
                    # 判断此单元格是否为空
                    if excel_data.cell(row=i, column=j).value is not None:
                        # 将单元格数据存入二维列表; .cell(x, y)获取x行y列的元素; .value获取元素的值
                        data[i - 2].append(excel_data.cell(row=i, column=j).value)
                    else:
                        data[i - 2].append([])
        return data
