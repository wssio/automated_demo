# -* coding:utf-8 -*-
"""
time   :  2021/3/07 17:53
author :  zhangzhilong
effect :  读取Excel文件并转换为代码执行
"""

# workbook()新建表格
# from openpyxl import workbook
# wb = workbook()
from openpyxl import load_workbook
from keyword_demo.base.bases import Base


class excel_keywords:
    @staticmethod
    def read_excel_keyword(drives, url, path):
        # 初始化基类
        drive = Base(drives)
        # 调用基类中的open_url方法打开网页
        drive.open_url(url)
        # load_workbook(文件名)打开已有表格，[工作簿名]获取工作簿
        excel_data = load_workbook(path[0])[path[1]]
        # range(x,y)从x到y-1的序列; .max_row获取工作簿的行数
        for i in range(1, int(str(excel_data.max_row)) + 1):
            # 排除表头
            if i != 1:
                # 定义空列表，用于存放从Excel中读取到的数据
                data = []
                # 定义空字符串，用于存放列表中的值转换成的基类语句
                command = ''
                # .max_column获取工作簿的列数
                for j in range(1, int(str(excel_data.max_column)) + 1):
                    # 判断Excel中此单元格是否为空
                    if excel_data.cell(row=i, column=j).value is not None:
                        # 如果不为空，则将元素值存入列表
                        data.append(excel_data.cell(row=i, column=j).value)
                    else:
                        # 如果为空，存入空字符串''，继续调用excel_data.cell(row=i, column=j).value则会存入None
                        data.append('')
                # 判断Excel中’操作‘的类型，根据类型选择对应操作
                if data[2] == "input":
                    # %s为占位符
                    command = "drive.%s ('%s','%s','%s','%s')" % (data[2], data[3], data[4], data[5], i)
                elif data[2] == "click":
                    # 这里判断Excel中此行是否有预期结果，如果预期不为空，则调用带断言的方法
                    if data[6] != '':
                        # 调用带断言的方法，传入i值，断言失败时记录第i行断言没有通过
                        command = "drive.%s('%s','%s','%s','%s') " % ((data[2]), data[3], data[4], data[6], i)
                    else:
                        # 调用不带断言的方法
                        command = "drive.%s('%s','%s') " % (data[2], data[3], data[4])
                elif data[2] == "sleep":
                    # 这里判断Excel中此行是否有预期结果，如果预期不为空，则调用带断言的方法
                    if data[6] != '':
                        # 调用带断言的方法，传入i值，断言失败时记录第i行断言没有通过
                        command = "drive.%s('%s','%s','%s') " % (data[2], data[5], data[6], i)
                    else:
                        # 调用不带断言的方法
                        command = "drive.%s('%s') " % (data[2], data[5])
                elif data[2] == "ctrl_c":
                    command = "drive.%s('%s','%s','%s','%s') " % (data[2], data[3], data[4], data[5], i)
                elif data[2] == "ctrl_v":
                    command = "drive.%s('%s','%s','%s') " % (data[2], data[3], data[4], i)
                elif data[2] == "handler":
                    # 这里判断Excel中此行是否有预期结果，如果预期不为空，则调用带断言的方法
                    if data[6] != '':
                        # 调用带断言的方法，传入i值，断言失败时记录第i行断言没有通过
                        command = "drive.%s('%s', '%s', '%s') " % (data[2], data[5], data[6], i)
                    else:
                        # 调用不带断言的方法
                        command = "drive.%s('%s') " % (data[2], data[5])
                else:
                    # 占位
                    pass
                # 执行command中存入的基类方法
                eval(command)
